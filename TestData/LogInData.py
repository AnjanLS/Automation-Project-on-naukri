import openpyxl


class LogInData:

    # test_LogIn_Data = [{"email": "arun.98@gmail.com", "password": "Arun@1998"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("D:/Anjan_L_S/python_testing/e-commerce/TestData/E-CommerceData.xls")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]